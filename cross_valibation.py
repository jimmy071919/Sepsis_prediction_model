import pandas as pd
import numpy as np
from sklearn.model_selection import StratifiedKFold, cross_validate
from sklearn.tree import DecisionTreeClassifier
from sklearn.svm import SVC
from sklearn.ensemble import RandomForestClassifier
from sklearn.neural_network import MLPClassifier
from sklearn.linear_model import LogisticRegression, SGDClassifier
from sklearn.metrics import make_scorer, precision_score, recall_score, f1_score, roc_auc_score
from sklearn.impute import SimpleImputer
import joblib
import os
import warnings
warnings.filterwarnings('ignore')

from load_data import load_data

class CrossValidationTrainer:
    """使用10-fold交叉驗證的模型訓練器"""
    
    def __init__(self, random_state=42, n_folds=10):
        """
        初始化交叉驗證訓練器
        
        Args:
            random_state: 隨機種子
            n_folds: 交叉驗證折數
        """
        self.random_state = random_state
        self.n_folds = n_folds
        self.models = self._initialize_models()
        self.cv_results = {}
        self.test_results = {}
        self.trained_models = {}  # 儲存訓練好的模型
        # 初始化數據填充器 (用於處理NaN值)
        self.imputer = SimpleImputer(strategy='median')
        # 初始化交叉驗證策略
        self.cv_strategy = StratifiedKFold(n_splits=n_folds, shuffle=True, random_state=random_state)
        
    def _initialize_models(self):
        """初始化模型"""
        models = {
            'DT': DecisionTreeClassifier(random_state=self.random_state, max_depth=10),
            'SVM': SVC(random_state=self.random_state, probability=True, kernel='rbf', C=1.0),
            'RF': RandomForestClassifier(random_state=self.random_state, n_estimators=50, max_depth=10),
            'ANN': MLPClassifier(random_state=self.random_state, hidden_layer_sizes=(50,), 
                                max_iter=300, early_stopping=True, validation_fraction=0.1),
            'LR': LogisticRegression(random_state=self.random_state, max_iter=1000, solver='liblinear'),
            'NN': MLPClassifier(random_state=self.random_state, hidden_layer_sizes=(100, 50, 25), 
                               max_iter=500, early_stopping=True, validation_fraction=0.1, alpha=0.01),
            'SGD': SGDClassifier(random_state=self.random_state, max_iter=1000, loss='log_loss', alpha=0.01)
        }
        return models
    
    def _get_scoring_metrics(self):
        """定義評估指標"""
        scoring = {
            'auc': 'roc_auc',
            'precision': make_scorer(precision_score, average='binary', zero_division=0),
            'recall': make_scorer(recall_score, average='binary', zero_division=0),
            'f1': make_scorer(f1_score, average='binary', zero_division=0)
        }
        return scoring
    
    def _custom_cross_validate(self, model, X, y, feature_type):
        """自定義交叉驗證，在每個fold內部正確處理NaN值"""
        from sklearn.base import clone
        
        cv_scores = {'auc': [], 'precision': [], 'recall': [], 'f1': []}
        
        for fold_idx, (train_idx, val_idx) in enumerate(self.cv_strategy.split(X, y)):
            print(f"    處理 Fold {fold_idx + 1}/{self.n_folds}...")
            
            # 分割當前fold的數據
            X_fold_train, X_fold_val = X[train_idx], X[val_idx]
            y_fold_train, y_fold_val = y[train_idx], y[val_idx]
            
            # 在當前fold的訓練集上處理NaN值
            if np.isnan(X_fold_train).any() or np.isnan(X_fold_val).any():
                fold_imputer = SimpleImputer(strategy='median')
                X_fold_train = fold_imputer.fit_transform(X_fold_train)
                X_fold_val = fold_imputer.transform(X_fold_val)
            
            # 訓練模型
            fold_model = clone(model)
            fold_model.fit(X_fold_train, y_fold_train)
            
            # 預測
            y_pred = fold_model.predict(X_fold_val)
            y_prob = fold_model.predict_proba(X_fold_val)[:, 1] if hasattr(fold_model, 'predict_proba') else fold_model.decision_function(X_fold_val)
            
            # 計算指標
            cv_scores['auc'].append(roc_auc_score(y_fold_val, y_prob))
            cv_scores['precision'].append(precision_score(y_fold_val, y_pred, average='binary', zero_division=0))
            cv_scores['recall'].append(recall_score(y_fold_val, y_pred, average='binary', zero_division=0))
            cv_scores['f1'].append(f1_score(y_fold_val, y_pred, average='binary', zero_division=0))
        
        # 轉換為numpy數組
        for metric in cv_scores:
            cv_scores[metric] = np.array(cv_scores[metric])
        
        return cv_scores
    
    def load_and_prepare_data(self, feature_type):
        """載入和準備指定特徵類型的數據"""
        try:
            # 載入結構化數據
            X_train_structured = np.load('structured_data_embedding/x_train_ax_scaled.npy')
            X_test_structured = np.load('structured_data_embedding/x_test_ax_scaled.npy')
            
            # 載入文本嵌入
            X_train_chief = np.load('unstructured_data_embedding/x_train_chief_embed.npy')
            X_test_chief = np.load('unstructured_data_embedding/x_test_chief_embed.npy')
            
            X_train_diagnosis = np.load('unstructured_data_embedding/x_train_diagnosis_embed.npy')
            X_test_diagnosis = np.load('unstructured_data_embedding/x_test_diagnosis_embed.npy')
            
            # 載入標籤
            y_train = np.load('answer_embedding/y_train.npy')
            y_test = np.load('answer_embedding/y_test.npy')
            
        except FileNotFoundError as e:
            raise FileNotFoundError(f"數據文件不存在: {e}")
        
        # 根據特徵類型組合訓練和測試數據
        if feature_type == 'a-x':
            X_train = X_train_structured
            X_test = X_test_structured
        elif feature_type == 'y':
            X_train = X_train_chief
            X_test = X_test_chief
        elif feature_type == 'z':
            X_train = X_train_diagnosis
            X_test = X_test_diagnosis
        elif feature_type == 'y-z':
            X_train = np.concatenate([X_train_chief, X_train_diagnosis], axis=1)
            X_test = np.concatenate([X_test_chief, X_test_diagnosis], axis=1)
        elif feature_type == 'a-y':
            X_train = np.concatenate([X_train_structured, X_train_chief], axis=1)
            X_test = np.concatenate([X_test_structured, X_test_chief], axis=1)
        elif feature_type == 'a-x,z':
            X_train = np.concatenate([X_train_structured, X_train_diagnosis], axis=1)
            X_test = np.concatenate([X_test_structured, X_test_diagnosis], axis=1)
        elif feature_type == 'a-z':
            X_train = np.concatenate([X_train_structured, X_train_chief, X_train_diagnosis], axis=1)
            X_test = np.concatenate([X_test_structured, X_test_chief, X_test_diagnosis], axis=1)
        else:
            raise ValueError(f"不支援的特徵類型: {feature_type}")
        
        # 注意：NaN值處理已移至交叉驗證內部，避免資料洩漏
        
        return X_train, X_test, y_train, y_test
    
    def perform_cross_validation(self, feature_type):
        """對指定特徵類型執行10-fold交叉驗證"""
        print(f"\\n=== 開始 {feature_type} 特徵的10-fold交叉驗證 ===")
        
        # 載入數據
        X_train, X_test, y_train, y_test = self.load_and_prepare_data(feature_type)
        
        print(f"訓練集特徵矩陣: {X_train.shape}")
        print(f"測試集特徵矩陣: {X_test.shape}")
        print(f"訓練集標籤: {y_train.shape}")
        print(f"測試集標籤: {y_test.shape}")
        
        cv_results = {}
        test_results = {}
        
        # 對每個模型進行交叉驗證
        for model_name, model in self.models.items():
            print(f"\\n正在進行 {model_name} 模型的10-fold交叉驗證...")
            
            try:
                # 檢查是否需要處理NaN值
                has_nan = feature_type in ['a-x', 'a-y', 'a-x,z', 'a-z'] and np.isnan(X_train).any()
                
                if has_nan:
                    # 使用自定義的交叉驗證來正確處理NaN值
                    cv_scores = self._custom_cross_validate(model, X_train, y_train, feature_type)
                else:
                    # 使用標準的交叉驗證
                    scoring = self._get_scoring_metrics()
                    cv_scores_dict = cross_validate(
                        model, X_train, y_train,
                        cv=self.cv_strategy,
                        scoring=scoring,
                        n_jobs=-1,
                        return_train_score=False
                    )
                    # 轉換格式以匹配自定義函數的輸出
                    cv_scores = {
                        'auc': cv_scores_dict['test_auc'],
                        'precision': cv_scores_dict['test_precision'],
                        'recall': cv_scores_dict['test_recall'],
                        'f1': cv_scores_dict['test_f1']
                    }
                
                # 計算交叉驗證平均分數和標準差
                cv_result = {
                    'AUC_mean': np.mean(cv_scores['auc']),
                    'AUC_std': np.std(cv_scores['auc']),
                    'precision_mean': np.mean(cv_scores['precision']),
                    'precision_std': np.std(cv_scores['precision']),
                    'recall_mean': np.mean(cv_scores['recall']),
                    'recall_std': np.std(cv_scores['recall']),
                    'f1_mean': np.mean(cv_scores['f1']),
                    'f1_std': np.std(cv_scores['f1'])
                }
                
                cv_results[model_name] = cv_result
                
                print(f"{model_name} 交叉驗證完成:")
                print(f"  AUC: {cv_result['AUC_mean']:.3f} (±{cv_result['AUC_std']:.3f})")
                print(f"  F1: {cv_result['f1_mean']:.3f} (±{cv_result['f1_std']:.3f})")
                
                # 在整個訓練集上訓練模型，然後在測試集上評估
                print(f"正在訓練 {model_name} 並在測試集上評估...")
                
                # 為最終模型訓練處理NaN值
                if has_nan:
                    fold_imputer = SimpleImputer(strategy='median')
                    X_train_final = fold_imputer.fit_transform(X_train)
                    X_test_final = fold_imputer.transform(X_test)
                else:
                    X_train_final = X_train
                    X_test_final = X_test
                
                model.fit(X_train_final, y_train)
                
                # 保存訓練好的模型
                if feature_type not in self.trained_models:
                    self.trained_models[feature_type] = {}
                self.trained_models[feature_type][model_name] = model
                
                # 在測試集上預測
                y_test_pred = model.predict(X_test_final)
                y_test_prob = model.predict_proba(X_test_final)[:, 1] if hasattr(model, 'predict_proba') else model.decision_function(X_test_final)
                
                # 計算測試集指標
                test_result = {
                    'AUC': roc_auc_score(y_test, y_test_prob),
                    'precision': precision_score(y_test, y_test_pred, average='binary', zero_division=0),
                    'recall': recall_score(y_test, y_test_pred, average='binary', zero_division=0),
                    'f1': f1_score(y_test, y_test_pred, average='binary', zero_division=0)
                }
                
                test_results[model_name] = test_result
                
                print(f"{model_name} 測試集結果:")
                print(f"  AUC: {test_result['AUC']:.3f}")
                print(f"  F1: {test_result['f1']:.3f}")
                
            except Exception as e:
                print(f"{model_name} 訓練失敗: {str(e)}")
                # 填充空結果
                empty_cv_result = {
                    'AUC_mean': 0, 'AUC_std': 0,
                    'precision_mean': 0, 'precision_std': 0,
                    'recall_mean': 0, 'recall_std': 0,
                    'f1_mean': 0, 'f1_std': 0
                }
                empty_test_result = {'AUC': 0, 'precision': 0, 'recall': 0, 'f1': 0}
                cv_results[model_name] = empty_cv_result
                test_results[model_name] = empty_test_result
        
        # 保存結果
        self.cv_results[feature_type] = cv_results
        self.test_results[feature_type] = test_results
        
        return cv_results, test_results
    
    def run_cross_validation_study(self):
        """執行完整的交叉驗證研究"""
        feature_types = ['a-x', 'y', 'z', 'y-z', 'a-y', 'a-x,z', 'a-z']
        
        print(f"=== 開始 {self.n_folds}-fold 交叉驗證研究 ===")
        
        for feature_type in feature_types:
            try:
                self.perform_cross_validation(feature_type)
            except Exception as e:
                print(f"特徵類型 {feature_type} 交叉驗證失敗: {str(e)}")
                continue
        
        # 生成總結報告
        self.create_summary_report()
    
    def create_summary_report(self):
        """創建交叉驗證結果總結報告"""
        print(f"\\n\\n=== {self.n_folds}-fold 交叉驗證結果總結 ===")
        
        for feature_type in self.cv_results:
            print(f"\\n=== {feature_type} 特徵 ===")
            
            # 交叉驗證結果
            print(f"\\n{self.n_folds}-fold 交叉驗證結果 (平均值 ± 標準差):")
            cv_df_data = []
            for model_name, results in self.cv_results[feature_type].items():
                cv_df_data.append({
                    'Model': model_name,
                    'AUC': f"{results['AUC_mean']:.3f} ± {results['AUC_std']:.3f}",
                    'Precision': f"{results['precision_mean']:.3f} ± {results['precision_std']:.3f}",
                    'Recall': f"{results['recall_mean']:.3f} ± {results['recall_std']:.3f}",
                    'F1': f"{results['f1_mean']:.3f} ± {results['f1_std']:.3f}"
                })
            
            cv_df = pd.DataFrame(cv_df_data)
            print(cv_df.to_string(index=False))
            
            # 測試集結果
            print(f"\\n測試集評估結果:")
            test_df_data = []
            for model_name, results in self.test_results[feature_type].items():
                test_df_data.append({
                    'Model': model_name,
                    'AUC': f"{results['AUC']:.3f}",
                    'Precision': f"{results['precision']:.3f}",
                    'Recall': f"{results['recall']:.3f}",
                    'F1': f"{results['f1']:.3f}"
                })
            
            test_df = pd.DataFrame(test_df_data)
            print(test_df.to_string(index=False))
    
    def save_results_to_xlsx(self):
        """將結果保存為Excel文件，每個特徵類型對應不同的工作表"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        from datetime import datetime
        
        # 創建Excel文件名
        excel_filename = 'result/cross_validation_results.xlsx'
        
        print(f"\n📊 正在保存結果到 Excel 檔案: {excel_filename}")
        
        # 創建工作簿
        wb = openpyxl.Workbook()
        # 移除默認工作表
        wb.remove(wb.active)
        
        # 定義樣式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # 創建摘要工作表
        summary_ws = wb.create_sheet("📊 摘要總覽")
        
        # 寫入摘要標題
        summary_ws['A1'] = f"敗血症預測模型 {self.n_folds}-fold 交叉驗證結果摘要"
        summary_ws['A1'].font = Font(size=16, bold=True)
        summary_ws['A2'] = f"生成時間: {datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')}"
        summary_ws['A2'].font = Font(size=12, italic=True)
        
        # 合併標題單元格
        summary_ws.merge_cells('A1:H1')
        summary_ws.merge_cells('A2:H2')
        
        summary_row = 4
        
        # 為每個特徵類型創建工作表並保存結果
        for feature_type in self.cv_results:
            print(f"   正在處理 {feature_type} 特徵...")
            
            # 創建交叉驗證結果工作表
            cv_ws_name = f"CV_{feature_type.replace('-', '_').replace(',', '_')}"
            cv_ws = wb.create_sheet(cv_ws_name)
            
            # 創建測試結果工作表
            test_ws_name = f"Test_{feature_type.replace('-', '_').replace(',', '_')}"
            test_ws = wb.create_sheet(test_ws_name)
            
            # 準備交叉驗證數據
            cv_data = []
            test_data = []
            
            for model_name in self.cv_results[feature_type]:
                # 交叉驗證結果
                cv_results = self.cv_results[feature_type][model_name]
                cv_data.append({
                    'Model': model_name,
                    'AUC_mean': round(cv_results['AUC_mean'], 4),
                    'AUC_std': round(cv_results['AUC_std'], 4),
                    'Precision_mean': round(cv_results['precision_mean'], 4),
                    'Precision_std': round(cv_results['precision_std'], 4),
                    'Recall_mean': round(cv_results['recall_mean'], 4),
                    'Recall_std': round(cv_results['recall_std'], 4),
                    'F1_mean': round(cv_results['f1_mean'], 4),
                    'F1_std': round(cv_results['f1_std'], 4)
                })
                
                # 測試集結果
                test_results = self.test_results[feature_type][model_name]
                test_data.append({
                    'Model': model_name,
                    'AUC': round(test_results['AUC'], 4),
                    'Precision': round(test_results['precision'], 4),
                    'Recall': round(test_results['recall'], 4),
                    'F1': round(test_results['f1'], 4)
                })
            
            # 創建DataFrame
            cv_df = pd.DataFrame(cv_data)
            test_df = pd.DataFrame(test_data)
            
            # 寫入交叉驗證工作表
            cv_ws['A1'] = f"{feature_type} 特徵 - {self.n_folds}-fold 交叉驗證結果"
            cv_ws['A1'].font = Font(size=14, bold=True)
            cv_ws.merge_cells('A1:I1')
            
            # 寫入CV數據
            for r in dataframe_to_rows(cv_df, index=False, header=True):
                cv_ws.append(r)
            
            # 寫入測試結果工作表
            test_ws['A1'] = f"{feature_type} 特徵 - 測試集評估結果"
            test_ws['A1'].font = Font(size=14, bold=True)
            test_ws.merge_cells('A1:E1')
            
            # 寫入測試數據
            for r in dataframe_to_rows(test_df, index=False, header=True):
                test_ws.append(r)
            
            # 格式化交叉驗證工作表
            self._format_worksheet(cv_ws, cv_df.shape[0] + 2, cv_df.shape[1], 
                                 header_font, header_fill, border, center_alignment)
            
            # 格式化測試結果工作表
            self._format_worksheet(test_ws, test_df.shape[0] + 2, test_df.shape[1], 
                                 header_font, header_fill, border, center_alignment)
            
            # 在摘要工作表中添加最佳結果
            summary_ws[f'A{summary_row}'] = f"{feature_type} 特徵最佳結果:"
            summary_ws[f'A{summary_row}'].font = Font(bold=True)
            summary_row += 1
            
            # 找出最佳AUC結果
            best_cv_auc = max(cv_data, key=lambda x: x['AUC_mean'])
            best_test_auc = max(test_data, key=lambda x: x['AUC'])
            
            summary_ws[f'B{summary_row}'] = f"交叉驗證最佳AUC: {best_cv_auc['Model']} ({best_cv_auc['AUC_mean']:.4f} ± {best_cv_auc['AUC_std']:.4f})"
            summary_row += 1
            summary_ws[f'B{summary_row}'] = f"測試集最佳AUC: {best_test_auc['Model']} ({best_test_auc['AUC']:.4f})"
            summary_row += 2
        
        # 調整摘要工作表列寬
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, 9):  # 假設最多8列
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            # 檢查這一列的所有單元格來計算最大長度
            for row in range(1, summary_row + 1):
                try:
                    cell = summary_ws.cell(row=row, column=col_idx)
                    if cell.value and hasattr(cell, 'value'):
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            # 設定列寬
            if max_length > 0:
                summary_ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # 保存Excel文件
        wb.save(excel_filename)
        
        print(f"✅ 結果已成功保存到 {excel_filename}")
        print(f"📋 包含 {len(self.cv_results)} 個特徵組合的詳細結果")
        print(f"📊 每個特徵組合都有獨立的交叉驗證和測試結果工作表")
    
    def _format_worksheet(self, ws, num_rows, num_cols, header_font, header_fill, border, center_alignment):
        """格式化工作表"""
        # 格式化標題行
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=2, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # 格式化數據行
        for row in range(3, num_rows + 1):
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.alignment = center_alignment
        
        # 調整列寬 - 修復合併單元格問題
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, num_cols + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            # 檢查這一列的所有單元格來計算最大長度
            for row in range(1, num_rows + 1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value and not isinstance(cell, type(ws.merged_cells)):
                    max_length = max(max_length, len(str(cell.value)))
            
            # 設定列寬
            ws.column_dimensions[column_letter].width = min(max_length + 2, 20)
    
    def save_trained_models(self):
        """保存所有訓練好的模型到檔案"""
        if not os.path.exists('models'):
            os.makedirs('models')
        
        print(f"\n🤖 正在保存訓練好的模型...")
        
        saved_count = 0
        for feature_type in self.trained_models:
            for model_name, model in self.trained_models[feature_type].items():
                # 創建模型檔名
                model_filename = f'models/{feature_type.replace("-", "_").replace(",", "_")}_{model_name}.pkl'
                
                # 保存模型
                joblib.dump(model, model_filename)
                saved_count += 1
                print(f"   ✅ {feature_type}-{model_name} 模型已保存: {model_filename}")
        
        print(f"\n📁 總共保存了 {saved_count} 個訓練好的模型到 models/ 目錄")
        print(f"💡 使用 joblib.load('模型路徑') 來載入模型進行預測")
        
        # 創建模型使用說明檔案
        self._create_model_usage_guide()
    
    def _create_model_usage_guide(self):
        """創建模型使用說明檔案"""
        guide_content = '''# 訓練模型使用說明

## 模型檔案說明

本目錄包含了敗血症預測模型的所有訓練好的模型檔案。

### 檔案命名規則
- 格式: `{特徵類型}_{模型名稱}.pkl`
- 特徵類型:
  - `a_x`: 僅結構化數據特徵
  - `y`: 僅主訴文本嵌入
  - `z`: 僅診斷文本嵌入
  - `a_y`: 結構化數據 + 主訴文本
  - `a_x_z`: 結構化數據 + 診斷文本
  - `a_z`: 結構化數據 + 主訴文本 + 診斷文本
- 模型名稱: DT(決策樹), SVM(支持向量機), RF(隨機森林), ANN(人工神經網路), LR(邏輯回歸), NN(神經網路), SGD(隨機梯度下降)

### 載入和使用模型

```python
import joblib
import numpy as np

# 載入模型
model = joblib.load('models/a_z_RF.pkl')  # 例如載入最佳組合的隨機森林模型

# 準備預測數據 (需要和訓練時相同的特徵順序和格式)
X_new = np.array([...])  # 新的病患數據

# 進行預測
y_pred = model.predict(X_new)  # 預測類別 (0: 無敗血症, 1: 有敗血症)
y_prob = model.predict_proba(X_new)[:, 1]  # 預測概率

print(f"預測結果: {y_pred[0]}")
print(f"敗血症概率: {y_prob[0]:.3f}")
```

### 注意事項
1. 使用模型前需確保輸入數據的預處理與訓練時一致
2. 結構化數據需要經過相同的標準化處理
3. 文本數據需要經過相同的BERT嵌入處理
4. 建議使用交叉驗證結果中表現最佳的模型組合
'''
        
        with open('models/README.md', 'w', encoding='utf-8') as f:
            f.write(guide_content)
        
        print(f"📝 模型使用說明已保存: models/README.md")
