"""
改良的交叉驗證訓練器
針對類別不平衡和高維特徵優化的版本
"""

import pandas as pd
import numpy as np
from sklearn.model_selection import StratifiedKFold, cross_validate
from sklearn.tree import DecisionTreeClassifier
from sklearn.svm import SVC
from sklearn.ensemble import RandomForestClassifier
from sklearn.neural_network import MLPClassifier
from sklearn.metrics import make_scorer, precision_score, recall_score, f1_score, roc_auc_score
from sklearn.impute import SimpleImputer
from imblearn.over_sampling import SMOTE
from imblearn.combine import SMOTETomek
import joblib
import os
import warnings
warnings.filterwarnings('ignore')

class OptimizedCrossValidationTrainer:
    """優化的交叉驗證訓練器"""
    
    def __init__(self, random_state=42, n_folds=10, use_smote=True):
        """
        初始化優化的交叉驗證訓練器
        
        Args:
            random_state: 隨機種子
            n_folds: 交叉驗證折數
            use_smote: 是否使用SMOTE處理類別不平衡
        """
        self.random_state = random_state
        self.n_folds = n_folds
        self.use_smote = use_smote
        self.models = self._initialize_optimized_models()
        self.cv_results = {}
        self.test_results = {}
        self.trained_models = {}
        # 初始化數據填充器
        self.imputer = SimpleImputer(strategy='median')
        # 初始化交叉驗證策略
        self.cv_strategy = StratifiedKFold(n_splits=n_folds, shuffle=True, random_state=random_state)
        # SMOTE初始化
        if self.use_smote:
            self.smote = SMOTETomek(random_state=random_state)
        
    def _initialize_optimized_models(self):
        """初始化優化的模型（針對類別不平衡調整參數）"""
        models = {
            'DT': DecisionTreeClassifier(
                random_state=self.random_state, 
                max_depth=10,
                class_weight='balanced'  # 處理類別不平衡
            ),
            'SVM': SVC(
                random_state=self.random_state, 
                probability=True, 
                kernel='rbf', 
                C=1.0,
                class_weight='balanced'  # 處理類別不平衡
            ),
            'RF': RandomForestClassifier(
                random_state=self.random_state, 
                n_estimators=100,  # 增加樹的數量
                max_depth=15,  # 稍微增加深度
                class_weight='balanced'  # 處理類別不平衡
            ),
            'CNN': MLPClassifier(
                random_state=self.random_state, 
                hidden_layer_sizes=(100, 50),  # 調整網路結構
                max_iter=500,  # 增加迭代次數
                early_stopping=True, 
                validation_fraction=0.1,
                alpha=0.001  # 正則化參數
            )
        }
        return models
    
    def _get_scoring_metrics(self):
        """定義評估指標"""
        scoring = {
            'auc': 'roc_auc',
            'precision': make_scorer(precision_score, average='binary', zero_division=0),
            'recall': make_scorer(recall_score, average='binary', zero_division=0),
            'f1': make_scorer(f1_score, average='binary', zero_division=0)
        }
        return scoring
    
    def load_and_prepare_optimized_data(self, feature_type):
        """載入和準備優化後的數據"""
        try:
            # 載入結構化數據
            X_train_structured = np.load('structured_data_embedding/x_train_ax_scaled.npy')
            X_test_structured = np.load('structured_data_embedding/x_test_ax_scaled.npy')
            
            # 載入優化後的文本嵌入（PCA降維後）
            X_train_chief = np.load('optimized_embeddings/x_train_chief_pca.npy')
            X_test_chief = np.load('optimized_embeddings/x_test_chief_pca.npy')
            
            X_train_diagnosis = np.load('optimized_embeddings/x_train_diagnosis_pca.npy')
            X_test_diagnosis = np.load('optimized_embeddings/x_test_diagnosis_pca.npy')
            
            # 載入標籤
            y_train = np.load('answer_embedding/y_train.npy')
            y_test = np.load('answer_embedding/y_test.npy')
            
        except FileNotFoundError as e:
            raise FileNotFoundError(f"優化後的數據文件不存在: {e}")
        
        # 根據特徵類型組合數據
        if feature_type == 'a-x':
            X_train = X_train_structured
            X_test = X_test_structured
        elif feature_type == 'y':
            X_train = X_train_chief
            X_test = X_test_chief
        elif feature_type == 'z':
            X_train = X_train_diagnosis
            X_test = X_test_diagnosis
        elif feature_type == 'a-y':
            X_train = np.concatenate([X_train_structured, X_train_chief], axis=1)
            X_test = np.concatenate([X_test_structured, X_test_chief], axis=1)
        elif feature_type == 'a-x,z':
            X_train = np.concatenate([X_train_structured, X_train_diagnosis], axis=1)
            X_test = np.concatenate([X_test_structured, X_test_diagnosis], axis=1)
        elif feature_type == 'a-z':
            X_train = np.concatenate([X_train_structured, X_train_chief, X_train_diagnosis], axis=1)
            X_test = np.concatenate([X_test_structured, X_test_chief, X_test_diagnosis], axis=1)
        else:
            raise ValueError(f"不支援的特徵類型: {feature_type}")
        
        print(f"   特徵類型 {feature_type} - 訓練集: {X_train.shape}, 測試集: {X_test.shape}")
        
        return X_train, X_test, y_train, y_test
    
    def _optimized_cross_validate(self, model, X, y, feature_type):
        """優化的交叉驗證，包含SMOTE處理"""
        from sklearn.base import clone
        
        cv_scores = {'auc': [], 'precision': [], 'recall': [], 'f1': []}
        
        for fold_idx, (train_idx, val_idx) in enumerate(self.cv_strategy.split(X, y)):
            print(f"    處理 Fold {fold_idx + 1}/{self.n_folds}...")
            
            # 分割當前fold的數據
            X_fold_train, X_fold_val = X[train_idx], X[val_idx]
            y_fold_train, y_fold_val = y[train_idx], y[val_idx]
            
            # 處理NaN值（如果包含結構化數據）
            if feature_type in ['a-x', 'a-y', 'a-x,z', 'a-z'] and np.isnan(X_fold_train).any():
                fold_imputer = SimpleImputer(strategy='median')
                X_fold_train = fold_imputer.fit_transform(X_fold_train)
                X_fold_val = fold_imputer.transform(X_fold_val)
            
            # 對純文本特徵使用SMOTE
            if self.use_smote and feature_type in ['y', 'z']:
                try:
                    X_fold_train, y_fold_train = self.smote.fit_resample(X_fold_train, y_fold_train)
                    print(f"      Fold {fold_idx + 1} SMOTE後數量: {len(y_fold_train)}")
                except Exception as e:
                    print(f"      Fold {fold_idx + 1} SMOTE失敗: {e}")
            
            # 訓練模型
            fold_model = clone(model)
            fold_model.fit(X_fold_train, y_fold_train)
            
            # 預測
            y_pred = fold_model.predict(X_fold_val)
            y_prob = fold_model.predict_proba(X_fold_val)[:, 1] if hasattr(fold_model, 'predict_proba') else fold_model.decision_function(X_fold_val)
            
            # 計算指標
            cv_scores['auc'].append(roc_auc_score(y_fold_val, y_prob))
            cv_scores['precision'].append(precision_score(y_fold_val, y_pred, average='binary', zero_division=0))
            cv_scores['recall'].append(recall_score(y_fold_val, y_pred, average='binary', zero_division=0))
            cv_scores['f1'].append(f1_score(y_fold_val, y_pred, average='binary', zero_division=0))
        
        # 轉換為numpy數組
        for metric in cv_scores:
            cv_scores[metric] = np.array(cv_scores[metric])
        
        return cv_scores
    
    def perform_optimized_cross_validation(self, feature_type):
        """執行優化的交叉驗證"""
        print(f"\\n=== 開始優化的 {feature_type} 特徵交叉驗證 ===")
        
        # 載入優化後的數據
        X_train, X_test, y_train, y_test = self.load_and_prepare_optimized_data(feature_type)
        
        cv_results = {}
        test_results = {}
        
        # 對每個模型進行交叉驗證
        for model_name, model in self.models.items():
            print(f"\\n正在進行 {model_name} 模型的優化交叉驗證...")
            
            try:
                # 執行優化的交叉驗證
                cv_scores = self._optimized_cross_validate(model, X_train, y_train, feature_type)
                
                # 計算交叉驗證結果
                cv_result = {
                    'AUC_mean': np.mean(cv_scores['auc']),
                    'AUC_std': np.std(cv_scores['auc']),
                    'precision_mean': np.mean(cv_scores['precision']),
                    'precision_std': np.std(cv_scores['precision']),
                    'recall_mean': np.mean(cv_scores['recall']),
                    'recall_std': np.std(cv_scores['recall']),
                    'f1_mean': np.mean(cv_scores['f1']),
                    'f1_std': np.std(cv_scores['f1'])
                }
                
                cv_results[model_name] = cv_result
                
                print(f"{model_name} 優化交叉驗證完成:")
                print(f"  AUC: {cv_result['AUC_mean']:.3f} (±{cv_result['AUC_std']:.3f})")
                print(f"  F1: {cv_result['f1_mean']:.3f} (±{cv_result['f1_std']:.3f})")
                print(f"  Recall: {cv_result['recall_mean']:.3f} (±{cv_result['recall_std']:.3f})")
                
                # 在測試集上評估
                print(f"正在訓練 {model_name} 並在測試集上評估...")
                
                # 準備最終訓練數據
                X_train_final = X_train.copy()
                X_test_final = X_test.copy()
                y_train_final = y_train.copy()
                
                # 處理NaN值
                if feature_type in ['a-x', 'a-y', 'a-x,z', 'a-z'] and np.isnan(X_train_final).any():
                    final_imputer = SimpleImputer(strategy='median')
                    X_train_final = final_imputer.fit_transform(X_train_final)
                    X_test_final = final_imputer.transform(X_test_final)
                
                # 對純文本特徵使用SMOTE
                if self.use_smote and feature_type in ['y', 'z']:
                    try:
                        X_train_final, y_train_final = self.smote.fit_resample(X_train_final, y_train_final)
                        print(f"  最終訓練集SMOTE後數量: {len(y_train_final)}")
                    except Exception as e:
                        print(f"  最終訓練集SMOTE失敗: {e}")
                
                # 訓練最終模型
                model.fit(X_train_final, y_train_final)
                
                # 保存模型
                if feature_type not in self.trained_models:
                    self.trained_models[feature_type] = {}
                self.trained_models[feature_type][model_name] = model
                
                # 測試集預測
                y_test_pred = model.predict(X_test_final)
                y_test_prob = model.predict_proba(X_test_final)[:, 1] if hasattr(model, 'predict_proba') else model.decision_function(X_test_final)
                
                # 計算測試集指標
                test_result = {
                    'AUC': roc_auc_score(y_test, y_test_prob),
                    'precision': precision_score(y_test, y_test_pred, average='binary', zero_division=0),
                    'recall': recall_score(y_test, y_test_pred, average='binary', zero_division=0),
                    'f1': f1_score(y_test, y_test_pred, average='binary', zero_division=0)
                }
                
                test_results[model_name] = test_result
                
                print(f"{model_name} 測試集結果:")
                print(f"  AUC: {test_result['AUC']:.3f}")
                print(f"  F1: {test_result['f1']:.3f}")
                print(f"  Recall: {test_result['recall']:.3f}")
                
            except Exception as e:
                print(f"{model_name} 訓練失敗: {str(e)}")
                # 填充空結果
                empty_cv_result = {
                    'AUC_mean': 0, 'AUC_std': 0,
                    'precision_mean': 0, 'precision_std': 0,
                    'recall_mean': 0, 'recall_std': 0,
                    'f1_mean': 0, 'f1_std': 0
                }
                empty_test_result = {'AUC': 0, 'precision': 0, 'recall': 0, 'f1': 0}
                cv_results[model_name] = empty_cv_result
                test_results[model_name] = empty_test_result
        
        # 保存結果
        self.cv_results[feature_type] = cv_results
        self.test_results[feature_type] = test_results
        
        return cv_results, test_results
    
    def run_optimized_cross_validation_study(self):
        """執行完整的優化交叉驗證研究"""
        feature_types = ['a-x', 'y', 'z', 'a-y', 'a-x,z', 'a-z']
        
        print(f"=== 開始優化的 {self.n_folds}-fold 交叉驗證研究 ===")
        if self.use_smote:
            print("📈 啟用SMOTE類別平衡處理")
        
        for feature_type in feature_types:
            try:
                self.perform_optimized_cross_validation(feature_type)
            except Exception as e:
                print(f"特徵類型 {feature_type} 優化交叉驗證失敗: {str(e)}")
                continue
        
        # 生成比較報告
        self.create_optimization_comparison()
    
    def create_optimization_comparison(self):
        """創建優化前後的比較報告"""
        print(f"\\n\\n=== 優化交叉驗證結果總結 ===")
        
        for feature_type in self.cv_results:
            print(f"\\n=== {feature_type} 特徵（優化版本） ===")
            
            # 交叉驗證結果
            print(f"\\n{self.n_folds}-fold 優化交叉驗證結果:")
            cv_df_data = []
            for model_name, results in self.cv_results[feature_type].items():
                cv_df_data.append({
                    'Model': model_name,
                    'AUC': f"{results['AUC_mean']:.3f} ± {results['AUC_std']:.3f}",
                    'Precision': f"{results['precision_mean']:.3f} ± {results['precision_std']:.3f}",
                    'Recall': f"{results['recall_mean']:.3f} ± {results['recall_std']:.3f}",
                    'F1': f"{results['f1_mean']:.3f} ± {results['f1_std']:.3f}"
                })
            
            cv_df = pd.DataFrame(cv_df_data)
            print(cv_df.to_string(index=False))
            
            # 測試集結果
            print(f"\\n測試集評估結果:")
            test_df_data = []
            for model_name, results in self.test_results[feature_type].items():
                test_df_data.append({
                    'Model': model_name,
                    'AUC': f"{results['AUC']:.3f}",
                    'Precision': f"{results['precision']:.3f}",
                    'Recall': f"{results['recall']:.3f}",
                    'F1': f"{results['f1']:.3f}"
                })
            
            test_df = pd.DataFrame(test_df_data)
            print(test_df.to_string(index=False))
    
    def save_optimized_results_to_xlsx(self):
        """保存優化結果到Excel"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from datetime import datetime
        
        excel_filename = 'result/optimized_cross_validation_results.xlsx'
        print(f"\\n📊 正在保存優化結果到: {excel_filename}")
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # 定義樣式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # 創建摘要工作表
        summary_ws = wb.create_sheet("📊 優化結果摘要")
        summary_ws['A1'] = f"敗血症預測模型優化版 {self.n_folds}-fold 交叉驗證結果"
        summary_ws['A1'].font = Font(size=16, bold=True)
        summary_ws['A2'] = f"生成時間: {datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')}"
        summary_ws['A3'] = f"優化措施: PCA降維 + 類別平衡處理"
        
        # 為每個特徵類型創建工作表
        for feature_type in self.cv_results:
            self._create_optimized_feature_sheet(wb, feature_type, header_font, header_fill, border, center_alignment)
        
        wb.save(excel_filename)
        print(f"✅ 優化結果已成功保存到 {excel_filename}")
    
    def _create_optimized_feature_sheet(self, wb, feature_type, header_font, header_fill, border, center_alignment):
        """為特定特徵類型創建工作表"""
        from openpyxl.styles import Font
        
        ws = wb.create_sheet(f"{feature_type}_優化版")
        
        # 交叉驗證結果
        ws['A1'] = f"{feature_type} 特徵組合 - 優化交叉驗證結果"
        ws['A1'].font = Font(size=14, bold=True)
        
        cv_headers = ['Model', 'AUC_Mean', 'AUC_Std', 'Precision_Mean', 'Precision_Std', 
                     'Recall_Mean', 'Recall_Std', 'F1_Mean', 'F1_Std']
        
        for col, header in enumerate(cv_headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # 填充交叉驗證數據
        row = 4
        for model_name, results in self.cv_results[feature_type].items():
            ws.cell(row=row, column=1, value=model_name)
            ws.cell(row=row, column=2, value=results['AUC_mean'])
            ws.cell(row=row, column=3, value=results['AUC_std'])
            ws.cell(row=row, column=4, value=results['precision_mean'])
            ws.cell(row=row, column=5, value=results['precision_std'])
            ws.cell(row=row, column=6, value=results['recall_mean'])
            ws.cell(row=row, column=7, value=results['recall_std'])
            ws.cell(row=row, column=8, value=results['f1_mean'])
            ws.cell(row=row, column=9, value=results['f1_std'])
            row += 1
        
        # 測試集結果  
        from openpyxl.styles import Font
        ws.cell(row=row+1, column=1, value="測試集結果").font = Font(size=12, bold=True)
        test_headers = ['Model', 'AUC', 'Precision', 'Recall', 'F1']
        
        for col, header in enumerate(test_headers, 1):
            cell = ws.cell(row=row+3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # 填充測試集數據
        test_row = row + 4
        for model_name, results in self.test_results[feature_type].items():
            ws.cell(row=test_row, column=1, value=model_name)
            ws.cell(row=test_row, column=2, value=results['AUC'])
            ws.cell(row=test_row, column=3, value=results['precision'])
            ws.cell(row=test_row, column=4, value=results['recall'])
            ws.cell(row=test_row, column=5, value=results['f1'])
            test_row += 1
    
    def save_optimized_models(self):
        """保存優化後的模型"""
        if not os.path.exists('optimized_models'):
            os.makedirs('optimized_models')
        
        print(f"\\n🤖 正在保存優化後的模型...")
        
        saved_count = 0
        for feature_type in self.trained_models:
            for model_name, model in self.trained_models[feature_type].items():
                model_filename = f'optimized_models/{feature_type.replace("-", "_").replace(",", "_")}_{model_name}_optimized.pkl'
                joblib.dump(model, model_filename)
                saved_count += 1
                print(f"   ✅ {feature_type}-{model_name} 優化模型已保存: {model_filename}")
        
        print(f"\\n📁 總共保存了 {saved_count} 個優化模型到 optimized_models/ 目錄")
        
        # 創建優化模型使用說明
        self._create_optimized_model_guide()
    
    def _create_optimized_model_guide(self):
        """創建優化模型使用說明"""
        guide_content = '''# 優化模型使用說明

## 優化內容

### 1. PCA降維
- 診斷文本嵌入: 768維 → 30維
- 主訴文本嵌入: 768維 → 30維
- 保留約70-80%的變異量

### 2. 類別平衡處理
- 對文本特徵使用SMOTE過採樣
- 所有模型使用class_weight='balanced'
- 針對類別不平衡問題優化

### 3. 模型參數調優
- Random Forest: 增加樹數量，調整深度
- SVM: 添加類別權重平衡
- Neural Network: 調整網路結構和正則化

## 預期改善效果

1. **文本特徵F1分數提升**: 特別是y和z特徵組合
2. **降低過擬合**: 通過PCA降維減少維度災難
3. **改善召回率**: 通過SMOTE和class_weight處理類別不平衡
4. **提高訓練效率**: 較低維度的特徵空間

## 使用方式

```python
import joblib
import numpy as np

# 載入優化模型
model = joblib.load('optimized_models/a_y_RF_optimized.pkl')

# 注意：輸入數據需要經過相同的預處理
# 1. 結構化數據需要標準化
# 2. 文本數據需要PCA降維到相同維度
```

## 重要提醒

使用優化模型前，確保：
1. 新數據經過相同的PCA轉換
2. 結構化數據使用相同的標準化器
3. 特徵順序與訓練時一致
'''
        
        with open('optimized_models/README.md', 'w', encoding='utf-8') as f:
            f.write(guide_content)
        
        print("📝 優化模型使用說明已保存: optimized_models/README.md")